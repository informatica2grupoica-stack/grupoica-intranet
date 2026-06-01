import re
import io
import urllib.parse
import time
import random
import requests
import json
from datetime import datetime
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from functools import lru_cache
try:
    from bs4 import BeautifulSoup
    BS4_DISPONIBLE = True
except ImportError:
    BS4_DISPONIBLE = False

PLAYWRIGHT_DISPONIBLE = False
PLAYWRIGHT_ERROR = "playwright reemplazado por selenium"

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options as _ChromeOptions
    from selenium.webdriver.chrome.service import Service as _ChromeService
    from selenium.webdriver.support.ui import WebDriverWait as _WebDriverWait
    from selenium.webdriver.support import expected_conditions as _EC
    from selenium.webdriver.common.by import By as _By
    SELENIUM_DISPONIBLE = True
    SELENIUM_ERROR = None
except Exception as _sel_err:
    SELENIUM_DISPONIBLE = False
    SELENIUM_ERROR = str(_sel_err)
try:
    import openpyxl
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

app = Flask(__name__)
CORS(app)

IVA = 1.19

cache_resultados = {}
CACHE_TTL = 300  # 5 minutos

HEADERS_BROWSER = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "es-CL,es;q=0.9",
}

# ─── Tiendas VTEX Chile ───────────────────────────────────────────────────────
# Desde IP real (notebook) todas funcionan — en Vercel algunas bloqueaban
VTEX_STORES = [
    {"dominio": "www.easy.cl",        "nombre": "Easy",        "prioridad": 10},
    {"dominio": "www.construmart.cl", "nombre": "Construmart", "prioridad": 10},
    {"dominio": "www.imperial.cl",    "nombre": "Imperial",    "prioridad": 9},
]

DOMINIOS_CHILE = {
    "sodimac":     {"dominio": "sodimac.cl",         "prioridad": 10},
    "easy":        {"dominio": "easy.cl",             "prioridad": 10},
    "construmart": {"dominio": "construmart.cl",      "prioridad": 10},
    "imperial":    {"dominio": "imperial.cl",         "prioridad": 9},
    "chilemat":    {"dominio": "chilemat.cl",         "prioridad": 9},
    "placacentro": {"dominio": "placacentro.cl",      "prioridad": 9},
    "cic":         {"dominio": "cic.cl",              "prioridad": 9},
    "aceroscmpc":  {"dominio": "aceroscmpc.cl",       "prioridad": 9},
    "mvm":         {"dominio": "mvm.cl",              "prioridad": 8},
    "cintac":      {"dominio": "cintac.cl",           "prioridad": 8},
    "sherwin":     {"dominio": "sherwin-williams.cl", "prioridad": 8},
    "sipa":        {"dominio": "sipa.cl",             "prioridad": 8},
    "seton":       {"dominio": "seton.cl",            "prioridad": 8},
    "prevenco":    {"dominio": "prevenco.cl",         "prioridad": 8},
    "mercadolibre":{"dominio": "mercadolibre.cl",     "prioridad": 6},
    "falabella":   {"dominio": "falabella.com",       "prioridad": 5},
    "paris":       {"dominio": "paris.cl",            "prioridad": 5},
    "ripley":      {"dominio": "ripley.cl",           "prioridad": 5},
}

INDICADORES_EXTRANJEROS = [
    "amazon.com", "ebay.com", "aliexpress", "wish.com",
    "walmart.com", "homedepot.com", "lowes.com",
    ".com.ar", ".com.mx", ".com.pe", ".com.co",
    "mercadolibre.com.ar", "mercadolibre.com.mx",
]

MONEDAS_EXTRANJERAS = re.compile(r'\b(USD|EUR|ARS|PEN|COP|MXN)\b', re.IGNORECASE)

CATEGORIAS_PRODUCTOS = {
    "madera": {
        "palabras_clave": ["madera", "pino", "mdf", "osb", "terciado", "plywood", "listón", "tablón", "tabla", "eucalipto"],
        "tiendas_prioritarias": ["construmart", "sodimac", "easy", "placacentro"],
        "queries_extra": ["madera {producto} Chile precio", "{producto} maderas precio CLP"],
        "unidades_relevantes": ["metro", "mt", "m2", "tabla", "unidad"],
    },
    "metal_acero": {
        "palabras_clave": ["fierro", "acero", "tubo", "tubular", "barra", "pletina", "pilar", "viga", "placa acero", "ángulo", "canal", "perfil"],
        "tiendas_prioritarias": ["aceroscmpc", "mvm", "cintac"],
        "queries_extra": ["{producto} acero Chile precio", "{producto} metalúrgica Chile"],
        "unidades_relevantes": ["kg", "metro", "barra", "mt"],
    },
    "cemento_hormigon": {
        "palabras_clave": ["cemento", "cal", "arena", "grava", "estabilizado", "hormigón", "concreto", "mortero"],
        "tiendas_prioritarias": ["chilemat", "construmart", "cic"],
        "queries_extra": ["{producto} precio Chile kg", "{producto} saco Chile ferretería"],
        "unidades_relevantes": ["saco", "kg", "m3", "bolsa"],
    },
    "ferreteria_general": {
        "palabras_clave": ["clavo", "tornillo", "perno", "malla", "soldadura", "alambre", "desmoldante", "tuerca", "remache"],
        "tiendas_prioritarias": ["sodimac", "easy", "imperial", "construmart"],
        "queries_extra": ["{producto} ferretería Chile precio", "{producto} precio Chile"],
        "unidades_relevantes": ["caja", "kg", "unidad", "paquete"],
    },
    "pintura_recubrimiento": {
        "palabras_clave": ["pintura", "anticorrosivo", "esmalte", "látex", "barniz", "semibrillo", "microesferas", "imprimante", "sellador"],
        "tiendas_prioritarias": ["sherwin", "sipa", "sodimac", "easy"],
        "queries_extra": ["{producto} pintura Chile litro precio", "{producto} Chile galón precio"],
        "unidades_relevantes": ["litro", "galón", "lt", "gl", "cuñete"],
    },
    "senaletica": {
        "palabras_clave": ["letrero", "señal", "tránsito", "paso cebra", "señalética", "delineador", "tachas", "tineta", "fastrack"],
        "tiendas_prioritarias": ["seton", "prevenco"],
        "queries_extra": ["{producto} señalética Chile precio", "{producto} tránsito Chile comprar"],
        "unidades_relevantes": ["unidad", "metro", "m2", "kit"],
    },
    "herramienta_medicion": {
        "palabras_clave": ["lienza", "rollo", "plomada", "nivel", "cinta", "metro"],
        "tiendas_prioritarias": ["sodimac", "easy", "imperial"],
        "queries_extra": ["{producto} Chile precio", "{producto} ferretería Chile"],
        "unidades_relevantes": ["unidad", "metro"],
    },
}


def get_cache_key(producto, limite):
    return f"{producto}_{limite}"


def limpiar_cache_expirado():
    ahora = time.time()
    for k in [k for k, v in cache_resultados.items() if ahora - v['timestamp'] > CACHE_TTL]:
        del cache_resultados[k]


def clasificar_producto(nombre: str) -> str:
    nombre_lower = nombre.lower()
    for categoria, datos in CATEGORIAS_PRODUCTOS.items():
        for palabra in datos["palabras_clave"]:
            if palabra in nombre_lower:
                return categoria
    return "ferreteria_general"


@lru_cache(maxsize=2000)
def normalizar(texto):
    if not texto:
        return ""
    texto = texto.lower()
    for orig, rep in {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n'}.items():
        texto = texto.replace(orig, rep)
    texto = re.sub(r'(\d+)\s*[xX×]\s*(\d+)', r'\1 x \2', texto)
    texto = re.sub(r'(\d+)\s*["\']\s*[xX×]\s*(\d+)\s*["\']', r'\1 x \2', texto)
    texto = re.sub(r'(\d+)\s*["\'](?![xX×])', r'\1 pulgadas', texto)
    texto = re.sub(r'(\d+)\s*[\']', r'\1 pies', texto)
    texto = re.sub(r'([a-z])\s+(\d)', r'\1\2', texto)
    texto = re.sub(r'(\d)\s+([a-z])', r'\1\2', texto)
    texto = re.sub(r'(\d+)\s+(\d+)\/(\d+)',
                   lambda m: str(float(m.group(1)) + float(m.group(2))/float(m.group(3))), texto)
    texto = re.sub(r'(\d+)\/(\d+)',
                   lambda m: str(float(m.group(1))/float(m.group(2))), texto)
    texto = re.sub(r'[^\w\s\/\-]', ' ', texto)
    return re.sub(r'\s+', ' ', texto).strip()


def extraer_medidas(texto: str) -> dict:
    medidas = {}
    t = texto.lower()
    dim3 = re.findall(r'(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)', t)
    if dim3: medidas['dim3'] = [float(d) for d in dim3[0]]
    dim2 = re.findall(r'(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)', t)
    if dim2: medidas['dim2'] = [float(d) for d in dim2[0]]
    frac = re.findall(r'(\d+)\s+(\d+)\/(\d+)|(\d+)\/(\d+)', t)
    for f in frac:
        medidas['fraccion'] = float(f[0]) + float(f[1])/float(f[2]) if f[0] else float(f[3])/float(f[4])
    pulg = re.findall(r'(\d+(?:\.\d+)?)\s*(?:"|pulgadas?|pulg\b|inch)', t)
    if pulg: medidas['pulgadas'] = float(pulg[0])
    mm = re.findall(r'(\d+(?:\.\d+)?)\s*mm', t)
    if mm: medidas['mm'] = [float(m) for m in mm]
    kg = re.findall(r'(\d+(?:\.\d+)?)\s*kg', t)
    if kg: medidas['kg'] = float(kg[0])
    lts = re.findall(r'(\d+(?:\.\d+)?)\s*(?:lt|lts|litros?|l\b)', t)
    if lts: medidas['litros'] = float(lts[0])
    mts = re.findall(r'(\d+(?:\.\d+)?)\s*(?:mt|mts|metros?|m\b)', t)
    if mts: medidas['metros'] = float(mts[0])
    return medidas


def medidas_a_texto(medidas: dict) -> str:
    partes = []
    if 'dim3' in medidas:
        v = medidas['dim3']; partes.append(f"{v[0]}x{v[1]}x{v[2]}")
    if 'dim2' in medidas:
        v = medidas['dim2']; partes.append(f"{v[0]}x{v[1]}")
    if 'fraccion' in medidas: partes.append(f"{medidas['fraccion']:.2f}\"")
    if 'pulgadas' in medidas: partes.append(f"{medidas['pulgadas']}\"")
    if 'mm' in medidas:
        mm = medidas['mm']; partes.append(f"{mm[0] if isinstance(mm, list) else mm}mm")
    if 'kg' in medidas: partes.append(f"{medidas['kg']}kg")
    if 'litros' in medidas: partes.append(f"{medidas['litros']}lt")
    if 'metros' in medidas: partes.append(f"{medidas['metros']}m")
    return ", ".join(partes) if partes else "sin medidas"


def comparar_medidas(medidas_b: dict, medidas_e: dict) -> float:
    if not medidas_b: return 0.5
    if not medidas_e: return 0.0
    coincidencias = total = 0
    for key, val_b in medidas_b.items():
        total += 1
        val_e = medidas_e.get(key)
        if val_e is None: continue
        if isinstance(val_b, list) and isinstance(val_e, list):
            if len(val_b) == len(val_e) and all(abs(a-b) < 0.5 for a, b in zip(sorted(val_b), sorted(val_e))):
                coincidencias += 1
        elif isinstance(val_b, (int, float)) and isinstance(val_e, (int, float)):
            if abs(val_b - val_e) <= max(0.1, val_b * 0.1): coincidencias += 1
        elif isinstance(val_b, list) and isinstance(val_e, (int, float)):
            if len(val_b) == 1 and abs(val_b[0] - val_e) < 0.5: coincidencias += 1
        elif isinstance(val_e, list) and isinstance(val_b, (int, float)):
            if len(val_e) == 1 and abs(val_b - val_e[0]) < 0.5: coincidencias += 1
    return coincidencias / total if total > 0 else 0.5


def extraer_especificaciones(texto: str) -> set:
    specs = set()
    t = normalizar(texto)
    for m in ['acero','hierro','fierro','pino','madera','cemento','cal','arena']:
        if m in t: specs.add(m)
    for tp in ['estriado','liso','bruto','estructural','tubular','cuadrado','rectangular',
               'redondo','galvanizado','cepillado','anticorrosivo','semibrillo']:
        if tp in t: specs.add(tp)
    if 'bruto' in t or 'sin cepillar' in t: specs.add('bruto')
    specs.update(re.findall(r'\b[a-z]+\d+\b', t))
    return specs


def calcular_concordancia(buscado: str, encontrado: str) -> int:
    b_norm = normalizar(buscado)
    e_norm = normalizar(encontrado)
    if not b_norm or not e_norm: return 0
    seq_ratio = SequenceMatcher(None, b_norm, e_norm).ratio()
    palabras_b = set(b_norm.split())
    palabras_e = set(e_norm.split())
    palabras_b_f = {p for p in palabras_b if len(p) > 2}
    jaccard = len(palabras_b_f & palabras_e) / len(palabras_b_f) if palabras_b_f else 0.5
    equiv = {'bruto':'sin cepillar','fierro':'acero','acero':'fierro',
             'anticorrosivo':'antioxidante','pino':'pino radiata','pino radiata':'pino'}
    b_eq, e_eq = b_norm, e_norm
    for orig, rep in equiv.items():
        b_eq = b_eq.replace(orig, rep)
        e_eq = e_eq.replace(orig, rep)
    seq_eq = SequenceMatcher(None, b_eq, e_eq).ratio()
    medidas_b = extraer_medidas(buscado)
    medidas_e = extraer_medidas(encontrado)
    med_score = comparar_medidas(medidas_b, medidas_e)
    bono = 0
    if medidas_b and medidas_e:
        bono = 15 if med_score >= 0.95 else (8 if med_score >= 0.7 else 0)
    specs_b = extraer_especificaciones(buscado)
    specs_e = extraer_especificaciones(encontrado)
    spec_match = len(specs_b & specs_e) / len(specs_b) if specs_b else 0.5
    score = (seq_ratio*0.15 + jaccard*0.30 + seq_eq*0.15 + med_score*0.25 + spec_match*0.15) * 100 + bono
    imp = {p for p in palabras_b_f if p in ['pino','madera','acero','fierro','cemento','pintura']}
    if imp: score -= len(imp - palabras_e) * 8
    if medidas_b and not medidas_e: score -= 15
    if len(palabras_b_f - palabras_e) > 3: score -= 10
    return round(min(100, max(0, score)))


def clasificar_concordancia(score: int):
    if score >= 90: return "exacta", "✅ Coincidencia exacta"
    if score >= 75: return "alta", "🟢 Alta coincidencia"
    if score >= 60: return "parcial", "🟡 Coincidencia parcial"
    if score >= 40: return "baja", "🟠 Baja coincidencia"
    return "nula", "🔴 Sin coincidencia"


def inferir_tipo_producto(nombre: str) -> dict:
    n = nombre.lower()
    return {
        "maquinaria_pesada": any(p in n for p in ["retroexcavadora","minicargador","grúa","compactador","pavimentadora"]),
        "herramienta_electrica": any(p in n for p in ["taladro","amoladora","sierra","esmeril","compresor","soldadora"]),
        "material_construccion": any(p in n for p in ["cemento","hormigón","arena","grava","madera","fierro","acero","tubo","placa","tabla","barra"]),
        "articulo_pequeno": any(p in n for p in ["clavo","tornillo","perno","tuerca","remache","tarugos"]),
        "pintura_quimico": any(p in n for p in ["pintura","anticorrosivo","barniz","esmalte","sellador","impermeabilizante"]),
        "senaletica_vial": any(p in n for p in ["letrero","señal","tránsito","paso cebra","tachas","delineador"]),
    }


def analizar_producto_buscado(nombre: str) -> dict:
    categoria = clasificar_producto(nombre)
    medidas = extraer_medidas(nombre)
    specs = list(extraer_especificaciones(nombre))
    nombre_norm = normalizar(nombre)
    palabras = [p for p in nombre_norm.split() if len(p) > 2]
    marcas = ["sherwin","sipa","gerdau","cintac","arauco","masisa","melon","stanley","bosch","dewalt","makita","hilti","sika"]
    return {
        "nombre_original": nombre,
        "nombre_normalizado": nombre_norm,
        "categoria": categoria,
        "palabras_clave": palabras,
        "medidas": {"tiene_medidas": bool(medidas), "detalle": medidas, "texto_legible": medidas_a_texto(medidas)},
        "especificaciones_tecnicas": specs,
        "unidades_relevantes": CATEGORIAS_PRODUCTOS.get(categoria, {}).get("unidades_relevantes", []),
        "es_accesorio": any(p in nombre.lower() for p in ["repuesto","accesorio","disco","carbón","estuche","funda"]),
        "marca_detectada": next((m for m in marcas if m in nombre.lower()), None),
        "tipo_producto": inferir_tipo_producto(nombre),
    }


def analizar_resultado_encontrado(resultado: dict, analisis_buscado: dict) -> dict:
    nombre = resultado.get("nombre", "")
    medidas_e = extraer_medidas(nombre)
    score = calcular_concordancia(analisis_buscado["nombre_original"], nombre)
    medidas_b = analisis_buscado["medidas"]["detalle"]
    conflicto = bool(medidas_b and medidas_e and comparar_medidas(medidas_b, medidas_e) < 0.5)
    palabras_b = set(analisis_buscado["palabras_clave"])
    palabras_e = set(normalizar(nombre).split())
    return {
        "medidas_encontradas": medidas_a_texto(medidas_e),
        "specs_encontradas": list(extraer_especificaciones(nombre)),
        "score_python": score,
        "nivel_python": clasificar_concordancia(score)[0],
        "palabras_comunes": list(palabras_b & palabras_e),
        "palabras_faltantes": list(palabras_b - palabras_e),
        "conflicto_medidas": conflicto,
    }


def limpiar_nombre(nombre: str) -> str:
    nombre = re.sub(r'\s*[\|–—]\s*.*$', '', nombre)
    nombre = re.sub(r'\s*MercadoLibre.*$', '', nombre, flags=re.IGNORECASE)
    nombre = re.sub(r'\s*Envío\s*(gratis|internacional|express).*$', '', nombre, flags=re.IGNORECASE)
    nombre = re.sub(r'\s*✓.*$', '', nombre)
    nombre = re.sub(r'\s*\(.*?\)$', '', nombre)
    return re.sub(r'\s+', ' ', nombre).strip()


def limpiar_precio(raw_price) -> int | None:
    precio_str = re.sub(r'[^\d]', '', str(raw_price))
    if not precio_str: return None
    precio = int(precio_str)
    return precio if 500 <= precio <= 500_000_000 else None


def es_resultado_chileno(item: dict) -> bool:
    url = item.get('url', item.get('link', '')).lower()
    tienda = item.get('tienda', item.get('source', '')).lower()
    for ind in INDICADORES_EXTRANJEROS:
        if ind in url or ind in tienda: return False
    if MONEDAS_EXTRANJERAS.search(str(item.get('price', item.get('precio_con_iva', '')))):
        return False
    if '.cl' in url: return True
    if any(v['dominio'] in url for v in DOMINIOS_CHILE.values()): return True
    if 'mercadolibre' in url and '/mlc' in url.lower(): return True
    return item.get('fuente', '') in ('mercadolibre_cl', 'vtex_direct', 'sodimac_direct')


def prioridad_tienda(url: str, tienda: str) -> int:
    url_l = url.lower(); tienda_l = tienda.lower()
    for nombre, datos in DOMINIOS_CHILE.items():
        if datos['dominio'] in url_l or nombre in tienda_l:
            return datos['prioridad']
    return 3


# ─── Google Shopping — Playwright (headless Chromium, ejecuta JS real) ────────

def _crear_driver_selenium():
    """Crea un Chrome headless con Selenium."""
    opts = _ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--lang=es-CL")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    # Selenium Manager descarga ChromeDriver automáticamente
    driver = webdriver.Chrome(options=opts)
    driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
    return driver


def buscar_google_shopping(producto: str, limite: int = 15):
    cache_key = get_cache_key(f"gshop_{producto}", limite)
    if cache_key in cache_resultados:
        return cache_resultados[cache_key]['data']
    if not SELENIUM_DISPONIBLE:
        print(f"  [GShop] Selenium no disponible: {SELENIUM_ERROR}")
        return []
    driver = None
    try:
        driver = _crear_driver_selenium()
        url = f"https://www.google.cl/search?q={urllib.parse.quote(producto + ' precio Chile')}&tbm=shop&hl=es&gl=cl&num=20"
        driver.get(url)

        # Esperar que carguen productos (máx 8s)
        try:
            _WebDriverWait(driver, 8).until(
                _EC.presence_of_element_located((_By.CSS_SELECTOR,
                    "div.sh-dgr__content, li.sh-dlr__list-result, div.KZmu8e, div[data-sh-gr]"))
            )
        except Exception:
            pass

        JS = """
        const res = [];
        const cards = document.querySelectorAll(
            'div.sh-dgr__content,li.sh-dlr__list-result,div.KZmu8e,div[data-sh-gr]'
        );
        for (const c of cards) {
            const h = c.querySelector('h3,h4,[role="heading"]');
            const nombre = h ? h.innerText.trim() : '';
            const pe = c.querySelector('b,strong,[class*="a8Pe"],[class*="e10tw"]');
            const precio = pe ? pe.innerText.trim() : (c.innerText.match(/\\$[\\d\\.]+/) || [''])[0];
            const a = c.querySelector('a');
            const link = a ? a.href : '';
            const se = c.querySelector('[class*="aULz"],[class*="IuHn"],[class*="E5oc"]');
            const tienda = se ? se.innerText.trim() : '';
            if (nombre && precio) res.push({nombre,precio,link,tienda});
        }
        if (!res.length) {
            const seen = new Set();
            for (const el of document.querySelectorAll('[aria-label]')) {
                const lbl = el.getAttribute('aria-label')||'';
                if (lbl.includes('$') && !seen.has(lbl)) { seen.add(lbl); res.push({nombre:lbl,precio:lbl,link:'',tienda:''}); }
            }
        }
        return res;
        """
        items = driver.execute_script(JS) or []
        resultados = []
        vistos = set()
        for item in items:
            nombre = limpiar_nombre(item.get("nombre", ""))
            if len(nombre) < 4 or nombre in vistos:
                continue
            precio = limpiar_precio(re.sub(r'[^\d]', '', item.get("precio", "")))
            if not precio:
                continue
            url_prod = item.get("link", "")
            tienda = (item.get("tienda", "") or "Google Shopping")[:40]
            if any(ind in url_prod.lower() for ind in INDICADORES_EXTRANJEROS):
                continue
            vistos.add(nombre)
            resultados.append({
                "tienda": tienda,
                "nombre": nombre[:150],
                "precio_con_iva": precio,
                "url": url_prod,
                "fuente": "google_shopping",
                "pais": "CL",
            })
            if len(resultados) >= limite:
                break
        if resultados:
            cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
        print(f"  [GShop] {len(resultados)} resultados (Selenium)")
        return resultados
    except Exception as e:
        print(f"  [GShop] Error Selenium: {e}")
        return []
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# ─── DuckDuckGo HTML (GET — maneja 200 y 202) ────────────────────────────────

def buscar_duckduckgo(producto: str, limite: int = 12):
    cache_key = get_cache_key(f"ddg_{producto}", limite)
    if cache_key in cache_resultados:
        return cache_resultados[cache_key]['data']
    if not BS4_DISPONIBLE:
        return []
    try:
        query = f"{producto} precio Chile comprar"
        r = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query, "kl": "cl-es"},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
                "Accept-Language": "es-CL,es;q=0.9",
                "Accept-Encoding": "gzip, deflate",
                "Referer": "https://duckduckgo.com/",
            },
            timeout=(5, 10),
            allow_redirects=True,
        )
        # DDG puede responder 200 o 202 — en ambos casos el HTML está en r.text
        if r.status_code not in (200, 202):
            print(f"  [DDG] HTTP {r.status_code}")
            return []

        soup = BeautifulSoup(r.text, "lxml")
        resultados = []

        for result in soup.select(".result")[:limite * 3]:
            a_tag = result.select_one(".result__a")
            snippet_tag = result.select_one(".result__snippet")
            if not a_tag:
                continue
            title = a_tag.get_text(strip=True)
            raw_href = a_tag.get("href", "")

            # Extraer URL real del redirect DDG (/l/?uddg=...)
            qs = urllib.parse.parse_qs(urllib.parse.urlparse(raw_href).query)
            url_res = qs.get("uddg", [raw_href])[0]
            if not url_res.startswith("http"):
                continue

            url_low = url_res.lower()
            if not any(x in url_low for x in [".cl", "mercadolibre"]):
                continue
            if any(ind in url_low for ind in INDICADORES_EXTRANJEROS):
                continue

            snippet = snippet_tag.get_text(strip=True) if snippet_tag else ""
            pm = re.search(r'\$\s*([\d\.,]{3,})', title + " " + snippet)
            if not pm:
                continue
            precio = limpiar_precio(re.sub(r'[^\d]', '', pm.group(1)))
            if not precio:
                continue

            domain_m = re.search(r'https?://(?:www\.)?([^/]+)', url_res)
            tienda = domain_m.group(1)[:40] if domain_m else "Web"

            resultados.append({
                "tienda": tienda,
                "nombre": limpiar_nombre(title)[:150],
                "precio_con_iva": precio,
                "url": url_res,
                "fuente": "duckduckgo",
                "pais": "CL",
            })
            if len(resultados) >= limite:
                break

        if resultados:
            cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
        print(f"  [DDG] {len(resultados)} resultados")
        return resultados
    except Exception as e:
        print(f"  [DDG] Error: {e}")
        return []


# ─── MercadoLibre Chile — scraping del website (API requiere OAuth) ───────────
# La API pública da 403 desde datacenter IPs. El website funciona desde IP residencial.

def buscar_mercadolibre(producto: str, limite: int = 12):
    cache_key = get_cache_key(f"meli_{producto}", limite)
    if cache_key in cache_resultados:
        return cache_resultados[cache_key]['data']
    if not BS4_DISPONIBLE:
        return []
    try:
        q = urllib.parse.quote(producto)
        r = requests.get(
            f"https://listado.mercadolibre.cl/{urllib.parse.quote(producto.replace(' ', '-'))}",
            params={"_NoIndex": "True"},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
                "Accept-Language": "es-CL,es;q=0.9",
                "Referer": "https://www.mercadolibre.cl/",
            },
            timeout=(5, 10),
        )
        if r.status_code not in (200, 301, 302):
            print(f"  [MeLi] HTTP {r.status_code}")
            return []

        soup = BeautifulSoup(r.text, "lxml")
        resultados = []

        # Selectores actuales MercadoLibre Chile (2024-2025)
        items = (soup.select("li.ui-search-layout__item")
                 or soup.select("div.ui-search-result__wrapper")
                 or soup.select(".andes-card"))

        for item in items:
            # Título
            title_el = (item.select_one("h2.ui-search-item__title")
                        or item.select_one(".poly-component__title")
                        or item.select_one("[class*='title']"))
            if not title_el:
                continue
            nombre = limpiar_nombre(title_el.get_text(strip=True))
            if len(nombre) < 4:
                continue

            # Precio — fracción entera
            precio_el = item.select_one("span.andes-money-amount__fraction")
            if not precio_el:
                # fallback: buscar cualquier monto en CLP
                pm = re.search(r'\$([\d\.]{3,})', item.get_text())
                if not pm:
                    continue
                precio = limpiar_precio(re.sub(r'[^\d]', '', pm.group(1)))
            else:
                precio = limpiar_precio(re.sub(r'[^\d]', '', precio_el.get_text(strip=True)))
            if not precio:
                continue

            # Link
            link_el = item.select_one("a[href*='mercadolibre.cl']") or item.select_one("a[href]")
            url = link_el.get("href", "").split("?")[0] if link_el else ""

            # Vendedor
            seller_el = item.select_one(".ui-search-official-store-label, [class*='seller']")
            tienda = seller_el.get_text(strip=True)[:40] if seller_el else "MercadoLibre CL"

            resultados.append({
                "tienda": tienda,
                "nombre": nombre[:150],
                "precio_con_iva": precio,
                "url": url,
                "fuente": "mercadolibre_cl",
                "pais": "CL",
            })
            if len(resultados) >= limite:
                break

        if resultados:
            cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
        print(f"  [MeLi] {len(resultados)} resultados")
        return resultados
    except Exception as e:
        print(f"  [MeLi] Error: {e}")
        return []


# ─── Sodimac Chile — API VTEX IO ─────────────────────────────────────────────

def buscar_sodimac(query: str, limite: int = 8):
    cache_key = get_cache_key(f"sodimac_{query}", limite)
    if cache_key in cache_resultados:
        return cache_resultados[cache_key]['data']
    if not BS4_DISPONIBLE:
        return []
    try:
        # Sodimac usa VTEX IO — endpoint de búsqueda con _q param
        urls_intentar = [
            f"https://www.sodimac.cl/api/catalog_system/pub/products/search?ft={urllib.parse.quote(query)}&_from=0&_to={limite-1}",
            f"https://www.sodimac.cl/sodimac-cl/search?q={urllib.parse.quote(query)}&map=ft&_from=0&_to={limite-1}",
        ]
        for url in urls_intentar:
            try:
                r = requests.get(url, headers={**HEADERS_BROWSER, "Accept": "application/json"}, timeout=(4, 8))
                if r.status_code == 200:
                    ct = r.headers.get("content-type", "")
                    if "json" in ct:
                        data = r.json()
                        if isinstance(data, list) and data:
                            resultados = []
                            for prod in data[:limite]:
                                nombre = limpiar_nombre(prod.get("productName", ""))
                                if len(nombre) < 4: continue
                                precio = None
                                link = prod.get("link", "")
                                for item in prod.get("items", []):
                                    for seller in item.get("sellers", []):
                                        offer = seller.get("commertialOffer", {})
                                        if offer.get("AvailableQuantity", 0) > 0:
                                            precio = offer.get("Price", 0)
                                            break
                                    if precio: break
                                if not precio or precio < 500: continue
                                if not link.startswith("http"):
                                    link = "https://www.sodimac.cl" + link
                                resultados.append({
                                    "tienda": "Sodimac",
                                    "nombre": nombre[:150],
                                    "precio_con_iva": round(precio),
                                    "url": link,
                                    "fuente": "vtex_direct",
                                    "pais": "CL",
                                })
                            if resultados:
                                cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
                                print(f"  [Sodimac] {len(resultados)} productos")
                                return resultados
            except Exception:
                continue
        print(f"  [Sodimac] 0 productos")
        return []
    except Exception as e:
        print(f"  [Sodimac] Error: {e}")
        return []


# ─── VTEX stores (Easy, Construmart, Imperial) ────────────────────────────────

def buscar_vtex(store: dict, query: str, limite: int = 8):
    """
    Busca en tiendas VTEX chilenas probando múltiples endpoints en orden.
    Easy, Construmart e Imperial migraron a VTEX IO — probamos varios paths.
    """
    cache_key = get_cache_key(f"vtex_{store['nombre']}_{query}", limite)
    if cache_key in cache_resultados:
        return cache_resultados[cache_key]['data']
    if not BS4_DISPONIBLE:
        return []

    dominio = store["dominio"]
    nombre_tienda = store["nombre"]
    q = urllib.parse.quote(query)

    # Endpoints a probar en orden — el primero que devuelva JSON válido gana
    endpoints = [
        # VTEX IO Intelligent Search (stores migradas)
        f"https://{dominio}/api/io/_v/api/intelligent-search/product_search?query={q}&page=1&count={limite}&sort=score_desc&locale=es-CL&hideUnavailableItems=true",
        # VTEX clásico
        f"https://{dominio}/api/catalog_system/pub/products/search?ft={q}&_from=0&_to={limite-1}",
        # VTEX IO con _q param
        f"https://{dominio}/api/io/search?_q={q}&map=ft&_from=0&_to={limite-1}",
    ]

    try:
        for url in endpoints:
            try:
                r = requests.get(url, headers={**HEADERS_BROWSER, "Accept": "application/json"}, timeout=(4, 8))
                if r.status_code != 200:
                    continue
                ct = r.headers.get("content-type", "")
                if "json" not in ct:
                    continue  # recibimos HTML — endpoint incorrecto
                data = r.json()
                productos = data if isinstance(data, list) else data.get("products", data.get("items", []))
                if not productos:
                    continue
                resultados = []
                for prod in productos[:limite]:
                    nombre = limpiar_nombre(prod.get("productName", prod.get("name", prod.get("ProductName", ""))))
                    if len(nombre) < 4:
                        continue
                    precio = None
                    link = prod.get("link", prod.get("url", prod.get("LinkId", "")))
                    for item in prod.get("items", [prod]):
                        for seller in item.get("sellers", []):
                            offer = seller.get("commertialOffer", {})
                            if offer.get("AvailableQuantity", 0) > 0:
                                precio = offer.get("Price", 0)
                                break
                        if precio:
                            break
                    if not precio:
                        precio = prod.get("price", prod.get("Price"))
                    if not precio or float(precio) < 500:
                        continue
                    if link and not link.startswith("http"):
                        link = f"https://{dominio}/{link}/p"
                    resultados.append({
                        "tienda": nombre_tienda,
                        "nombre": nombre[:150],
                        "precio_con_iva": round(float(precio)),
                        "url": link,
                        "fuente": "vtex_direct",
                        "pais": "CL",
                    })
                if resultados:
                    cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
                    print(f"  [{nombre_tienda}] {len(resultados)} productos (via {url.split('/')[4]})")
                    return resultados
            except Exception:
                continue

        # Fallback: scraping HTML del buscador de la tienda
        try:
            r = requests.get(
                f"https://{dominio}/buscador",
                params={"q": query},
                headers=HEADERS_BROWSER,
                timeout=(4, 10),
            )
            if r.status_code == 200 and BS4_DISPONIBLE:
                soup = BeautifulSoup(r.text, "lxml")
                resultados = []
                for card in soup.select("[data-product-id], .product-summary, .vtex-product-summary"):
                    nombre_el = card.select_one(".product-summary-name, h3, [class*='name']")
                    precio_el = card.select_one(".product-selling-price, [class*='price'], .sellingPrice")
                    if not nombre_el or not precio_el:
                        continue
                    nombre = limpiar_nombre(nombre_el.get_text(strip=True))
                    precio = limpiar_precio(re.sub(r'[^\d]', '', precio_el.get_text(strip=True)))
                    link_el = card.select_one("a[href]")
                    link = link_el.get("href", "") if link_el else ""
                    if not link.startswith("http"):
                        link = f"https://{dominio}{link}"
                    if nombre and precio:
                        resultados.append({
                            "tienda": nombre_tienda,
                            "nombre": nombre[:150],
                            "precio_con_iva": precio,
                            "url": link,
                            "fuente": "vtex_direct",
                            "pais": "CL",
                        })
                    if len(resultados) >= limite:
                        break
                if resultados:
                    cache_resultados[cache_key] = {"data": resultados, "timestamp": time.time()}
                    print(f"  [{nombre_tienda}] {len(resultados)} productos (HTML fallback)")
                    return resultados
        except Exception:
            pass

        print(f"  [{nombre_tienda}] 0 productos")
        return []
    except Exception as e:
        print(f"  [{nombre_tienda}] Error: {e}")
        return []


# ─── Generador de queries ─────────────────────────────────────────────────────

def generar_queries(producto: str, categoria: str) -> list:
    queries = [producto]
    prod_norm = re.sub(r'(\d+)\s*"', r'\1 pulgadas', producto)
    prod_norm = re.sub(r'(\d+)\s*\'', r'\1 pies', prod_norm)
    if prod_norm != producto: queries.append(prod_norm)
    terminos_extra = {
        "madera":             ["madera Chile"],
        "metal_acero":        ["acero Chile"],
        "cemento_hormigon":   ["saco Chile"],
        "senaletica":         ["señalética Chile"],
        "pintura_recubrimiento": ["galón Chile"],
    }
    for termino in terminos_extra.get(categoria, []):
        queries.append(f"{producto} {termino}")
    return list(dict.fromkeys(queries))[:3]


# ─── BÚSQUEDA PRINCIPAL ───────────────────────────────────────────────────────

def realizar_busqueda(producto: str, limite: int = 15, conversion: str = "unidad"):
    print(f"\n  🔍 Producto: {producto}")
    categoria = clasificar_producto(producto)
    print(f"  📂 Categoría: {categoria}")
    limpiar_cache_expirado()

    resultados = []
    urls_vistas = set()

    def agregar(nuevos):
        for r in nuevos:
            url = r.get('url', '')
            clave = url or normalizar(r.get('nombre', ''))
            if clave and clave not in urls_vistas:
                urls_vistas.add(clave)
                resultados.append(r)

    # ── Query VTEX: solo palabras clave sin números puros ni stop words ────────
    PALABRAS_IGNORAR = {'con', 'para', 'por', 'los', 'las', 'del', 'una', 'uno',
                        'marca', 'similar', 'unidades', 'unidad', 'medidas', 'varias',
                        'resistente', 'acabado', 'interior', 'exterior', 'brillante'}
    palabras_clave = [
        w for w in producto.split()
        if not re.match(r'^[\d\.,xX×"\'\/\-]+$', w)
        and len(w) > 2
        and w.lower() not in PALABRAS_IGNORAR
    ]
    # VTEX falla con acentos — normalizar antes de enviar
    def _sin_acentos(t):
        for a, b in {'á':'a','é':'e','í':'i','ó':'o','ú':'u','Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','ñ':'n','Ñ':'N'}.items():
            t = t.replace(a, b)
        return t
    query_vtex = _sin_acentos(' '.join(palabras_clave[:5]))

    print(f"  📡 MeLi + GShop + DDG + Sodimac + VTEX × {len(VTEX_STORES)} (VTEX: '{query_vtex}')")

    # IMPORTANTE: NO usar "with" — espera todos los threads al salir aunque haya timeout
    ex = ThreadPoolExecutor(max_workers=12)
    try:
        futures = {
            ex.submit(buscar_mercadolibre, producto, 12): "MercadoLibre",
            ex.submit(buscar_google_shopping, producto, limite): "Google Shopping",
            ex.submit(buscar_duckduckgo, producto, 12): "DuckDuckGo",
            ex.submit(buscar_sodimac, query_vtex, 8): "Sodimac",
        }
        for store in VTEX_STORES:
            futures[ex.submit(buscar_vtex, store, query_vtex, 8)] = store["nombre"]

        try:
            for future in as_completed(futures, timeout=18):
                nombre_f = futures[future]
                try:
                    nuevos = future.result()
                    agregar(nuevos)
                    print(f"  ✅ {nombre_f}: {len(nuevos)} | Total: {len(resultados)}")
                except Exception as e:
                    print(f"  ❌ {nombre_f}: {e}")
        except Exception:
            print(f"  ⚠️ Timeout 18s — continuando con {len(resultados)} resultados parciales")
    finally:
        try:
            ex.shutdown(wait=False, cancel_futures=True)
        except TypeError:
            ex.shutdown(wait=False)

    # ── Fase 2: variación de query VTEX si hay pocos resultados ──────────────
    if len(resultados) < 5:
        # Intentar query más corta (solo 3 palabras)
        query_corta = ' '.join(palabras_clave[:3])
        if query_corta and query_corta != query_vtex:
            print(f"  📡 Fase 2: VTEX query corta '{query_corta}'...")
            ex2 = ThreadPoolExecutor(max_workers=5)
            try:
                futures2 = {ex2.submit(buscar_vtex, store, query_corta, 6): store["nombre"]
                            for store in VTEX_STORES}
                try:
                    for future in as_completed(futures2, timeout=12):
                        try:
                            agregar(future.result())
                        except Exception:
                            pass
                except Exception:
                    pass
            finally:
                try:
                    ex2.shutdown(wait=False, cancel_futures=True)
                except TypeError:
                    ex2.shutdown(wait=False)

    if not resultados:
        return [], {}

    # ── Scoring y ranking ─────────────────────────────────────────────────────
    analisis_buscado = analizar_producto_buscado(producto)
    conv_lower = conversion.lower() if conversion else "unidad"
    for r in resultados:
        ar = analizar_resultado_encontrado(r, analisis_buscado)
        score = ar["score_python"]
        nombre_r = r.get("nombre", "").lower()
        if conv_lower not in ("unidad", "und", "un", ""):
            if detectar_unidad_resultado(nombre_r, conv_lower):
                score = min(100, score + 8)
        nivel, etiqueta = clasificar_concordancia(score)
        r.update({
            "score": score,
            "nivel_concordancia": nivel,
            "etiqueta_concordancia": etiqueta,
            "prioridad_tienda": prioridad_tienda(r.get("url", ""), r.get("tienda", "")),
            "categoria": categoria,
            "conversion": conv_lower,
            "precio_neto": round(r["precio_con_iva"] / IVA),
            "precio_formateado": f"${r['precio_con_iva']:,.0f}".replace(",", "."),
            "_analisis": ar,
        })

    resultados.sort(key=lambda x: (-x["score"], -x["prioridad_tienda"], x["precio_con_iva"]))
    filtrados = [r for r in resultados if r["score"] >= 10] or resultados
    return filtrados[:limite], analisis_buscado


# ─── ENDPOINT ─────────────────────────────────────────────────────────────────

UNIDADES_VALIDAS = {
    "kg":     ["kg", "kilo", "kilos", "kilogramo", "/kg", "por kg"],
    "m3":     ["m3", "m³", "metro cubico", "metro cúbico", "/m3", "por m3"],
    "m2":     ["m2", "m²", "metro cuadrado", "/m2"],
    "galon":  ["galón", "galon", "gal", "gl", "/gl", "1 gl", "1 galón"],
    "litro":  ["litro", "lt", "lts", "/lt", "1 lt", "1 litro"],
    "tira":   ["tira", "tira 6m", "/tira"],
    "caja":   ["caja", "por caja", "/caja"],
    "pack":   ["pack", "paquete", "por pack"],
    "tineta": ["tineta", "/tineta"],
    "metro":  ["metro", "mt", "/mt", "por metro"],
    "rollo":  ["rollo", "por rollo"],
    "unidad": [],
}


def detectar_unidad_resultado(nombre: str, conversion: str) -> bool:
    if not conversion or conversion.lower() in ("unidad", "und", "un", ""):
        return True
    nombre_lower = nombre.lower()
    conv_lower = conversion.lower()
    for key, indicadores in UNIDADES_VALIDAS.items():
        if conv_lower in (key, *indicadores):
            return any(ind in nombre_lower for ind in indicadores) if indicadores else True
    return True


@app.route("/python/busqueda-robusta", methods=["GET"])
def busqueda_robusta():
    producto = request.args.get("producto", "").strip()
    numero_item = request.args.get("numero", "")
    minimo_requerido = int(request.args.get("minimo", 15))
    force_refresh = request.args.get("force", "").lower() == "true"
    conversion = request.args.get("conversion", "unidad").strip().lower()

    print(f"\n{'='*60}\n🔍 [{numero_item}] {producto}\n{'='*60}")

    if not producto:
        return jsonify({"numero_item": numero_item, "producto": producto, "resultados": [],
                        "total_encontrados": 0, "suficientes": False, "deficit": minimo_requerido,
                        "categoria": "desconocida", "analisis_producto": {}})

    if force_refresh:
        for k in [k for k in cache_resultados if producto.lower() in k.lower()]:
            del cache_resultados[k]
        print(f"  🔄 Cache limpiado para: {producto}")

    resultados, analisis_buscado = realizar_busqueda(producto, minimo_requerido * 2, conversion=conversion)

    resultados_formateados = []
    for r in resultados:
        ar = r.pop("_analisis", {})
        resultados_formateados.append({
            "tienda": r.get("tienda", ""),
            "nombre": r.get("nombre", ""),
            "precio_valor": r.get("precio_con_iva", 0),
            "precio_neto": r.get("precio_neto", 0),
            "precio_formateado": r.get("precio_formateado", "Consultar"),
            "link": r.get("url", ""),
            "canal": r.get("fuente", "web"),
            "pais": "CL",
            "busqueda_original": producto,
            "score": r.get("score", 0),
            "nivel_concordancia": r.get("nivel_concordancia", ""),
            "etiqueta_concordancia": r.get("etiqueta_concordancia", ""),
            "categoria": r.get("categoria", ""),
            "prioridad_tienda": r.get("prioridad_tienda", 3),
            "medidas_encontradas": ar.get("medidas_encontradas", ""),
            "specs_encontradas": ar.get("specs_encontradas", []),
            "palabras_comunes": ar.get("palabras_comunes", []),
            "palabras_faltantes": ar.get("palabras_faltantes", []),
            "conflicto_medidas": ar.get("conflicto_medidas", False),
            "conversion": r.get("conversion", "unidad"),
        })

    tiene_suficientes = len(resultados_formateados) >= minimo_requerido
    mejor = resultados_formateados[0] if resultados_formateados else None

    print(f"\n📊 TOTAL: {len(resultados_formateados)} resultados chilenos")
    print(f"✅ Suficientes: {tiene_suficientes}")
    if mejor:
        print(f"🏆 Mejor: {mejor['tienda']} → ${mejor['precio_valor']:,} ({mejor['score']}%)")
    print("=" * 60)

    return jsonify({
        "numero_item": numero_item,
        "producto": producto,
        "categoria": clasificar_producto(producto),
        "resultados": resultados_formateados,
        "total_encontrados": len(resultados_formateados),
        "suficientes": tiene_suficientes,
        "deficit": max(0, minimo_requerido - len(resultados_formateados)),
        "pais_busqueda": "CL",
        "analisis_producto": analisis_buscado,
    })


@app.route("/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok",
        "pais": "Chile 🇨🇱",
        "cache_size": len(cache_resultados),
        "fuentes": ["Google Shopping (Selenium)", "MercadoLibre", "DuckDuckGo", "Sodimac"] + [s["nombre"] for s in VTEX_STORES],
        "version": "7.0"
    })


@app.route("/python/diagnostico", methods=["GET"])
def diagnostico():
    resultado = {}

    # Test MercadoLibre — API pública, debería funcionar siempre
    try:
        r_meli = requests.get(
            "https://api.mercadolibre.com/sites/MLC/search",
            params={"q": "tornillo acero", "limit": 5},
            headers={"Accept": "application/json"},
            timeout=8,
        )
        raw = r_meli.json() if r_meli.status_code == 200 else {}
        raw_items = raw.get("results", [])
        items = buscar_mercadolibre("tornillo acero", 5)
        resultado["mercadolibre"] = {
            "ok": len(items) > 0,
            "http": r_meli.status_code,
            "raw_count": len(raw_items),
            "first_id": raw_items[0].get("id", "") if raw_items else "",
            "resultados": len(items),
            "muestra": items[0].get("nombre", "") if items else "",
        }
    except Exception as e:
        resultado["mercadolibre"] = {"ok": False, "error": str(e)}

    # Test Google Shopping — captura HTTP status real
    try:
        r = requests.get(
            "https://www.google.cl/search",
            params={"q": "tornillo precio Chile", "tbm": "shop", "hl": "es", "gl": "cl"},
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"},
            timeout=8,
        )
        items = buscar_google_shopping("tornillo", 3)
        resultado["google_shopping"] = {
            "ok": len(items) > 0, "http": r.status_code,
            "html_len": len(r.text), "resultados": len(items),
            "muestra": items[0].get("nombre", "") if items else "",
        }
    except Exception as e:
        resultado["google_shopping"] = {"ok": False, "error": str(e)}

    # Test DuckDuckGo — captura HTTP status real
    try:
        r = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": "tornillo precio Chile", "kl": "cl-es"},
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"},
            timeout=8,
        )
        items = buscar_duckduckgo("tornillo", 3)
        resultado["duckduckgo"] = {
            "ok": len(items) > 0, "http": r.status_code,
            "html_len": len(r.text), "resultados": len(items),
            "muestra": items[0].get("nombre", "") if items else "",
        }
    except Exception as e:
        resultado["duckduckgo"] = {"ok": False, "error": str(e)}

    # Test Sodimac — captura HTTP status real
    try:
        r = requests.get(
            "https://www.sodimac.cl/sodimac-cl/search",
            params={"Ntt": "tornillo", "jsonContent": "true"},
            headers={**HEADERS_BROWSER, "Accept": "application/json"},
            timeout=8,
        )
        items = buscar_sodimac("tornillo", 3)
        resultado["sodimac"] = {
            "ok": len(items) > 0, "http": r.status_code,
            "body_preview": r.text[:200],
            "resultados": len(items),
            "muestra": items[0].get("nombre", "") if items else "",
        }
    except Exception as e:
        resultado["sodimac"] = {"ok": False, "error": str(e)}

    # Test VTEX — prueba account subdomains y endpoints alternativos
    vtex_tests = [
        ("easy_vtex_account",    "https://easy.vtexcommercestable.com.br/api/catalog_system/pub/products/search?ft=tornillo&_from=0&_to=2"),
        ("construmart_vtex_acct","https://construmart.vtexcommercestable.com.br/api/catalog_system/pub/products/search?ft=tornillo&_from=0&_to=2"),
        ("easy_s_endpoint",      "https://www.easy.cl/s?q=tornillo&map=ft&_from=0&_to=2"),
        ("easy_search_json",     "https://www.easy.cl/api/catalog_system/pub/products/search?fq=ft:tornillo&_from=0&_to=2"),
    ]
    for label, url in vtex_tests:
        try:
            r = requests.get(url, headers={**HEADERS_BROWSER, "Accept": "application/json"}, timeout=7)
            ct = r.headers.get("content-type", "")
            try:
                body = r.json()
                n = len(body) if isinstance(body, list) else len(body.get("products", body.get("items", [])))
            except Exception:
                body = {}
                n = 0
            resultado[label] = {
                "http": r.status_code,
                "content_type": ct[:60],
                "resultados": n,
                "body_preview": r.text[:150],
            }
        except Exception as e:
            resultado[label] = {"error": str(e)[:100]}

    # Test MercadoLibre website (no API)
    try:
        r_web = requests.get(
            "https://listado.mercadolibre.cl/tornillo-acero",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36", "Accept-Language": "es-CL"},
            timeout=8,
        )
        items_web = buscar_mercadolibre("tornillo acero", 3)
        resultado["mercadolibre_web"] = {
            "ok": len(items_web) > 0,
            "http": r_web.status_code,
            "html_len": len(r_web.text),
            "resultados": len(items_web),
            "muestra": items_web[0].get("nombre", "") if items_web else "",
        }
    except Exception as e:
        resultado["mercadolibre_web"] = {"ok": False, "error": str(e)}

    # Info del Python que corre Flask
    import sys as _sys, subprocess as _sub
    resultado["python_info"] = {
        "executable": _sys.executable,
        "version": _sys.version[:30],
    }
    try:
        pip_list = _sub.run([_sys.executable, "-m", "pip", "show", "playwright"],
                             stdout=_sub.PIPE, stderr=_sub.PIPE, text=True, timeout=5)
        resultado["python_info"]["playwright_pip"] = pip_list.stdout[:200] if pip_list.returncode == 0 else "no encontrado"
    except Exception as ep:
        resultado["python_info"]["playwright_pip"] = str(ep)[:100]

    # Test Playwright directo
    try:
        if not SELENIUM_DISPONIBLE:
            resultado["selenium"] = {"ok": False, "error": f"selenium no disponible: {SELENIUM_ERROR}"}
        else:
            driver = None
            try:
                driver = _crear_driver_selenium()
                driver.get("https://www.google.cl/search?q=tornillo+acero+precio+Chile&tbm=shop&hl=es&gl=cl")
                import time as _t; _t.sleep(4)
                title = driver.title
                html_len = len(driver.page_source)
                n_cards = driver.execute_script("return document.querySelectorAll('div.sh-dgr__content,li.sh-dlr__list-result,div.KZmu8e,div[data-sh-gr]').length")
                precios = driver.execute_script("return (document.body.innerText.match(/\\$[\\d\\.]{3,}/g)||[]).slice(0,5)")
                resultado["selenium"] = {"ok": True, "title": title[:60], "html_len": html_len, "cards": n_cards, "precios": precios}
            except Exception as e:
                resultado["selenium"] = {"ok": False, "error": str(e)[:200]}
            finally:
                if driver:
                    try: driver.quit()
                    except: pass
    except Exception as e:
        resultado["playwright"] = {"ok": False, "error": str(e)[:200]}

    alguna_ok = any(v.get("ok") for v in resultado.values())
    return jsonify({
        "estado_general": "operativo" if alguna_ok else "diagnostico_detallado",
        "fuentes": resultado,
        "server_ip": request.environ.get("HTTP_X_FORWARDED_FOR", request.remote_addr),
    })


@app.route("/python/debug-html", methods=["GET"])
def debug_html():
    """Devuelve fragmento del HTML real para identificar selectores CSS."""
    fuente = request.args.get("fuente", "gshop")
    q = request.args.get("q", "tornillo acero")

    if fuente == "gshop":
        r = requests.get(
            "https://www.google.cl/search",
            params={"q": q + " precio Chile", "tbm": "shop", "hl": "es", "gl": "cl", "num": 10},
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                     "Accept-Language": "es-CL,es;q=0.9"},
            timeout=10,
        )
        html = r.text
        # Extraer clases CSS únicas del HTML para identificar estructura
        import re as _re
        clases = _re.findall(r'class="([^"]{5,60})"', html)
        clases_unicas = list(dict.fromkeys(clases))[:80]
        # Buscar divs con precio
        fragmentos_precio = _re.findall(r'.{0,100}\$[\d\.]{3,}.{0,100}', html)[:10]
        return jsonify({
            "fuente": "google_shopping",
            "http": r.status_code,
            "html_len": len(html),
            "clases_css": clases_unicas,
            "fragmentos_con_precio": fragmentos_precio[:8],
            "html_inicio": html[5000:7000],
        })

    elif fuente == "meli":
        r = requests.get(
            f"https://listado.mercadolibre.cl/{urllib.parse.quote(q.replace(' ','-'))}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                     "Accept-Language": "es-CL,es;q=0.9"},
            timeout=10,
        )
        html = r.text
        import re as _re
        clases = _re.findall(r'class="([^"]{5,60})"', html)
        clases_unicas = list(dict.fromkeys(clases))[:80]
        fragmentos_precio = _re.findall(r'.{0,100}\$[\d\.]{3,}.{0,100}', html)[:10]
        return jsonify({
            "fuente": "mercadolibre_web",
            "http": r.status_code,
            "html_len": len(html),
            "clases_css": clases_unicas,
            "fragmentos_con_precio": fragmentos_precio[:8],
            "html_inicio": html[3000:5000],
        })

    return jsonify({"error": "fuente inválida — usa ?fuente=gshop o ?fuente=meli"})


@app.route("/python/cache/clear", methods=["POST"])
def clear_cache():
    cache_resultados.clear()
    return jsonify({"status": "ok", "mensaje": "Caché limpiado"})


@app.route("/python/exportar-costeo", methods=["POST"])
def exportar_costeo():
    if not OPENPYXL_DISPONIBLE:
        return jsonify({"error": "openpyxl no disponible en el servidor"}), 500
    if "archivo" not in request.files:
        return jsonify({"error": "No se envió el archivo Excel"}), 400

    archivo = request.files["archivo"]
    try:
        datos = json.loads(request.form.get("datos", "{}"))
    except Exception:
        return jsonify({"error": "Datos JSON inválidos"}), 400

    sheet_name = datos.get("sheet", "COSTEO")
    items_data = datos.get("items", [])
    if not items_data:
        return jsonify({"error": "No hay ítems con resultados"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()))
    except Exception as e:
        return jsonify({"error": f"No se pudo abrir el Excel: {e}"}), 400

    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row_idx = None
    col_item = col_valor_civa = col_link1 = None

    for row in ws.iter_rows(min_row=1, max_row=25):
        headers = {}
        for cell in row:
            h = str(cell.value or "").upper().strip()
            if h:
                headers[h] = cell.column
        if "ITEM" in headers:
            header_row_idx = row[0].row
            col_item = headers.get("ITEM")
            for h, c in headers.items():
                if "VALOR" in h and "IVA" in h:
                    col_valor_civa = c
                    break
            for h, c in sorted(headers.items(), key=lambda x: x[1]):
                if h.startswith("LINK"):
                    col_link1 = c
                    break
            break

    if not header_row_idx or not col_item:
        return jsonify({"error": "No se encontró columna ITEM en la pestaña seleccionada"}), 400
    if not col_valor_civa:
        return jsonify({"error": "No se encontró columna VALOR C/IVA"}), 400

    max_col = ws.max_column
    col_tienda_extra = max_col + 1
    col_match_extra = max_col + 2
    ws.cell(row=header_row_idx, column=col_tienda_extra).value = "TIENDA COTIZADA"
    ws.cell(row=header_row_idx, column=col_match_extra).value = "% COINCIDENCIA"

    precios = {str(i["numero"]): i for i in items_data}
    filled = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        cell_item = row[col_item - 1]
        if cell_item.value is None:
            continue
        item_str = str(cell_item.value).strip()
        if not item_str.isdigit():
            continue
        dato = precios.get(item_str)
        if not dato or not dato.get("precio"):
            continue

        r_num = cell_item.row
        ws.cell(row=r_num, column=col_valor_civa).value = dato["precio"]
        if col_link1 and dato.get("link"):
            ws.cell(row=r_num, column=col_link1).value = dato["link"]
        ws.cell(row=r_num, column=col_tienda_extra).value = dato.get("tienda", "")
        ws.cell(row=r_num, column=col_match_extra).value = f"{dato.get('match', 0)}%"
        filled += 1

    if filled == 0:
        return jsonify({"error": "No se encontraron ítems con resultados para exportar"}), 400

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.now().strftime("%Y-%m-%d")
    print(f"  📊 Export COSTEO: {filled} ítems rellenados en '{sheet_name}'")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"COSTEO_cotizado_{fecha}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    print("=" * 60)
    print("🚀 BUSCADOR CHILE — GRUPO ICA v7.0 — Google Shopping con Playwright")
    print("   FUENTES PRIMARIAS:")
    print("   • Google Shopping Chile (scraper tbm=shop — IP real)")
    print("   • DuckDuckGo Chile (scraper HTML — cualquier IP)")
    print("   FUENTES DIRECTAS (precios exactos):")
    for s in VTEX_STORES:
        print(f"   • {s['nombre']} (VTEX directo)")
    print("   ARQUITECTURA: ThreadPoolExecutor — shutdown(wait=False)")
    print("=" * 60)
    try:
        from waitress import serve
        print("✅ Waitress activo en http://0.0.0.0:5000")
        serve(app, host="0.0.0.0", port=5000, threads=16)
    except ImportError:
        print("⚠️  Waitress no instalado, usando Flask dev server")
        app.run(host="0.0.0.0", port=5000, debug=True)
